from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from .models import Doctor, Patient, Prescription, passwordHasher, emailHasher, Appointment, Medicine, timeofday, doctorlogo
from django.db.models import Count, Q
from django.contrib.auth.decorators import login_required
from .forms import AppointmentSet, AppointmentSetForm, AppointmentForm
from datetime import datetime, time, timedelta, date
import time
from django.utils import timezone
from django.shortcuts import render
import threading
import sys, os
# import pyautogui
import openpyxl
from django.db import connections, IntegrityError
from django.core.exceptions import ValidationError
from django.apps import apps
import requests
import zipfile
from django.conf import settings
# from WPP_Whatsapp import Create, PlaywrightSafeThread
from weasyprint import HTML, CSS
import pdfkit
from django.template.loader import render_to_string, get_template
from jinja2 import Template
import asyncio
from django.core.serializers import serialize
import json
from django.forms.models import model_to_dict
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import Doctor
from django.core.mail import send_mail
from django.conf import settings
import random
from .forms import ForgotPasswordForm, VerifyOTPForm, ResetPasswordForm
from django.views.decorators.csrf import csrf_exempt

if ('runserver' in sys.argv):
    from .Whatsapptestfile import whatsappApi, openWhatsapp, whatsappApiEdit, whatsappMedia, whatsappApiDoc
    # import Whatsapptestfile

renderedHTML = " "
def openwhatsapp(request):
    # openWhatsapp()
    openWhatsapp.wp()
    # return render(request, 'HealthCentre/adminPortal.html')
    return HttpResponseRedirect(reverse("admin"))

def closewhatsapp(request):
    openWhatsapp.closewp()
    # return render(request, 'HealthCentre/adminPortal.html')
    return HttpResponseRedirect(reverse("admin"))

def updateExcel(request):
    # while True:
        xlPath = os.curdir #"D:\Dental-Software-Backup\Dental-Software"
        allfilesinpath = os.listdir(xlPath)
        xlFile = [file for file in allfilesinpath if file.lower().startswith('databasetables.xlsx')]
        docName = request.session['Name']
        if not xlFile:
            workbook = openpyxl.Workbook()
            for sheetIndex, model in enumerate(apps.get_models()):
                if model.__name__ in [ 'Patient', 'Prescription', 'Appointment']:

                    if model.__name__ == 'Prescription':
                        excludedColumns = ['timestamp', 'doctor', 'patient']
                        worksheet = workbook.create_sheet(title='sheet2')
                        filterColumnName = 'prescribingDoctor'
                        
                    elif model.__name__ == 'Appointment':
                        excludedColumns = ['time', 'date', 'AppointmentTimeStamp']
                        worksheet = workbook.create_sheet(title='sheet3')
                        filterColumnName = 'appointmentdoctor'
                    else:
                        excludedColumns = []
                        worksheet = workbook.create_sheet(title='sheet1')
                        filterColumnName = 'doctorname'
                        excludedColumns = ['id', 'doctorid', 'emailHash', 'doctorname']

                    allColumns = [field.name for field in model._meta.fields]
                    includedColumns = [col for col in allColumns if col not in excludedColumns]
                    
                    with connections['default'].cursor() as cursor:
                        cursor.execute(f'SELECT {",".join(includedColumns)} FROM {model._meta.db_table} WHERE {filterColumnName} = %s', [docName])
                        rows = cursor.fetchall()
                        
                    headers = includedColumns
                    worksheet.append(headers)

                    for row in rows:
                        worksheet.append(row)
            try : 
                workbook.save('databasetables.xlsx')
                workbook.close()
            except PermissionError:
                time.sleep(1)
        else:
            try:
                workbookExisting = openpyxl.load_workbook('databasetables.xlsx')
            except (zipfile.BadZipfile) as exception:
                time.sleep(1)
                os.remove('databasetables.xlsx')
                updateExcel()
           
            for sheetIndex, model in enumerate(apps.get_models()):
                if model.__name__ in [ 'Patient', 'Prescription', 'Appointment']:
                    
                    if model.__name__ == 'Prescription':
                        excludedColumns = ['timestamp', 'doctor', 'patient']
                        worksheetExisting = workbookExisting.get_sheet_by_name('sheet2')
                        filterColumnName = 'prescribingDoctor'
                        
                    elif model.__name__ == 'Appointment':
                        excludedColumns = ['time', 'date', 'AppointmentTimeStamp']
                        worksheetExisting = workbookExisting.get_sheet_by_name('sheet3')
                        filterColumnName = 'appointmentdoctor'
                        
                    else:
                        excludedColumns = []
                        worksheetExisting = workbookExisting.get_sheet_by_name('sheet1')
                        filterColumnName = 'doctorname'
                        excludedColumns = ['id', 'doctorid', 'emailHash', 'doctorname']
                        
                    worksheetExisting.delete_rows(1, worksheetExisting.max_row)
                    allColumns = [field.name for field in model._meta.fields]
                    includedColumns = [col for col in allColumns if col not in excludedColumns]
                    
                    with connections['default'].cursor() as cursor:
                        cursor.execute(f'SELECT {",".join(includedColumns)} FROM {model._meta.db_table} WHERE {filterColumnName} = %s', [docName])
                        rows = cursor.fetchall()
                    headers = includedColumns
                    worksheetExisting.append(headers)

                    for row in rows:
                        worksheetExisting.append(row)
            try:
                workbookExisting.save('databasetables.xlsx')
                workbookExisting.close()
            except PermissionError:
                time.sleep(1)
        openExcel = open('databasetables.xlsx', 'rb')
        response = HttpResponse( openExcel.read(), content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename = databasetables.xlsx'
        return response


def uploadExcel(request):

    if request.method == 'GET':
        return HttpResponseRedirect(reverse('patMed'))
    
    if request.method == 'POST':
        excel_file = request.FILES['excel']
        # you may put validations here to check extension or file size
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.get_sheet_by_name('sheet1')
        for row in worksheet.iter_rows(min_row=2, values_only= True):
            name, address, contactNumber, email, age, sex = row
            existingPatient = Patient.objects.filter(name = name)
            for existingDetails in existingPatient:
                doctname = request.session['Name']
                doctpk = Doctor.objects.get(name = doctname)
                if existingDetails.name == name or existingDetails.rollNumber == age or existingDetails.email == email or existingDetails.passwordHash == sex or existingDetails.address == address or existingDetails.contactNumber == contactNumber or existingDetails.doctorname == doctname:
                    emailHash = emailHasher(email)
                    patient = Patient(name = name,rollNumber = age, email = email, passwordHash = sex, address = address, 
                                    contactNumber = contactNumber, emailHash = emailHash, doctorname = doctname, doctorid = doctpk)
                    patient.save()
            if not existingPatient :
                    doctorname = request.session['Name']
                    doctorpk = Doctor.objects.get(name = doctorname)
                    # Encrypting email to store inside database
                    emailHash = emailHasher(email)
                    patient = Patient(name = name,rollNumber = age, email = email, passwordHash = sex, address = address, 
                                    contactNumber = contactNumber, emailHash = emailHash, doctorname = doctorname, doctorid = doctorpk)
                    patient.save()
            
        response = render(request,"HealthCentre/medicinePatientPortal.html")
        return HttpResponseRedirect(reverse('patMed'))


def index(request):
    """ Function for displaying main page of website. """
    if request.method == 'GET':
        currentDate = datetime.now().date()
        curTime = datetime.now().time()
        docName = request.session['Name']
        # currentAppointments = Appointment.objects.filter(date__gte=currentDate,time__gt= curTime ,appointmentdoctor = docName).order_by('date').order_by('time')
        currentAppointments = Appointment.objects.filter(date__gte=currentDate,time__gt=curTime,appointmentdoctor=docName).order_by('date', 'time')
        nextFirstAppoint = None
        nextAppointmentData = []
        # for currentAppointment in currentAppointments:
        #     appointDate = currentAppointment.date
        #     appointTime = currentAppointment.time
        #     appointPatient = currentAppointment.appointmentpatient
        #     appointNotes = currentAppointment.notes
        if currentAppointments:
            nextFirstAppoint = currentAppointments.first()
            nextAppointmentData = currentAppointments[1:]
            status = "there are few more appointments today.."
        else:
            status = "there are no more appointments today.."
            
        context = {
            'nextFirstAppoint' : nextFirstAppoint,
            'curAppoints' : nextAppointmentData,
            'status' : status,
            # 'appointDate' : appointDate,
            # 'appointTime' : appointTime,
            # 'appointPatient' : appointPatient,
            # 'appointNotes' : appointNotes,
        }
        
        return render(request, "HealthCentre/index.html", context)
    
@csrf_exempt    
def updateDashboard(request):
    # if request.GET.get('currentAppointment') == None:
    if request.method =='POST':
        appointmntStatus = request.POST.get('status')
        appointmntId = request.POST.get('id')
        appointmentPat = Appointment.objects.get(pk=appointmntId)

        if appointmentPat.status == False:
            try:
                appointPat = Appointment.objects.get(pk=appointmntId)
                appointPat.status = True

                if appointPat.status == True : 
                    appointPat.status = True
                    appointPat.save()
                    return render(request, "HealthCentre/index.html", {'appointment': appointPat})
            except Appointment.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Appointment not found'})

    currentDate = datetime.now().date()
    curTime = datetime.now().time()
    docName = request.session['Name']
    currentAppointments = Appointment.objects.filter(date__gte=currentDate, time__gte= curTime, appointmentdoctor = docName).order_by('time')
    incompleteAppointments = Appointment.objects.filter(date__lte=currentDate, time__lte= curTime, appointmentdoctor = docName,status=False).order_by('time')
    completedAppointments = Appointment.objects.filter(date__lte=currentDate, time__lte= curTime, appointmentdoctor = docName,status=True).order_by('time')
    nextAppointmentData = []
    # for currentAppointment in currentAppointments:
    #     appointDate = currentAppointment.date
    #     appointTime = currentAppointment.time
    #     appointPatient = currentAppointment.appointmentpatient
    #     appointNotes = currentAppointment.notes
    #     nextAppointmentData.append({
    #         'appointDate' : appointDate,
    #         'appointTime' : appointTime,
    #         'appointPatient' : appointPatient,
    #         'appointNotes' : appointNotes,
            
    #     })
    if currentAppointments:
        
        nextFirstAppoint = currentAppointments.first()
        nextAppointmentData = serialize('json' ,currentAppointments[1:])
        status = "There are few more appointments today.."
        nextFirstAppointData = serialize('json', [nextFirstAppoint])[1:-1] if nextFirstAppoint else None
        data = {
            'lastIncompleteAppoint' : None,
            'nextFirstAppoint' : json.loads(nextFirstAppointData),
            'curAppoints' : json.loads(nextAppointmentData),
            'status' : status
        }
        
    else:    
        
        status = "There are no more appointments today.!."
        data = {
            'lastIncompleteAppoint' : None,
            'nextFirstAppoint' : None,
            'curAppoints' : None,
            'status' : status
        }
    if incompleteAppointments:
        latestIncompleteAppoint = incompleteAppointments.first()
        
        lastIncompleteAppoint = serialize('json',incompleteAppointments)
        
        data = {
            'lastIncompleteAppoint' : json.loads(lastIncompleteAppoint),
            'nextFirstAppoint' : None,
            'curAppoints' : None,
            'status' : None
        }
        
        


    # print(data)
    # data = dashAppointData
    # Editing response headers so as to ignore cached versions of pages
    # response = render(request,"HealthCentre/index.html")
    return JsonResponse(data)
# return render(request, "HealthCentre/index.html", context)
        # return render(request,"HealthCentre/index.html")

def register(request):
    """ Function for registering a student into the portal. """

    # If the user wants the page to get displayed

    if request.method == "GET":
        # Editing response headers so as to ignore cached versions of pages

        response =  render(request,"HealthCentre/registrationPortal.html")

        return responseHeadersModifier(response)
    
    # If the user wants to submit his/her information

    elif request.method == "POST":
        # Extracting the user information from the post request
        userFirstName = request.POST["userFirstNam"]
        userLastName = request.POST["userLastName"]
        userEmail = request.POST["userEmail"]
        userRollNo = request.POST["userRollNo"]
        userAddress = request.POST["userAddress"]
        userContactNo = request.POST["userContactNo"]
        userPassword = request.POST["userPassword"]
        
        userType = request.POST['userType']
        if userType == 'patient':
            userConfirmPassword = userPassword
        elif userType == 'doctor':
            userConfirmPassword = request.POST["userConfirmPassword"]
        # If both the passwords match
        if userPassword == userConfirmPassword:

            name = userFirstName + " " + userLastName

            

            # handleSubmit(request)
            # def handleSubmit(request):
            
            
            # Creating a patient object and saving insdie the database if patient is selected 
            
            if userType == 'patient':
                # patient = Patient(rollNumber=request.POST['rollNumber'])
                # Encrypting password to store inside database
                passwordHash = userPassword
                doctorname = request.session['Name']
                doctorpk = Doctor.objects.get(name = doctorname)
                # Encrypting email to store inside database
                emailHash = emailHasher(userEmail)
                try:
                    patient = Patient(name = name,rollNumber = userRollNo, email = userEmail, passwordHash = passwordHash, address = userAddress, 
                                    contactNumber = userContactNo, emailHash = emailHash, doctorname = doctorname, doctorid = doctorpk)
                    patient.save()
                except IntegrityError:
                    
                    message = "Patient Name or Email already exists!, Use a different Name or email"
                    context = {
                                "message": message
                            }

                    # Editing response headers so as to ignore cached versions of pages
                    response = render(request,"HealthCentre/registrationPortal.html",context)
                    return responseHeadersModifier(response)
            # Creating a patient object and saving insdie the database if patient is selected
            elif userType == 'doctor':
                clinicName = request.POST['clinicName']
                educationQualification = request.POST['educationQualification']
                
                passwordHash = passwordHasher(userPassword)
                emailHash = emailHasher(userEmail)
                try:
                    doctor = Doctor(name = name, specialization= userRollNo, email = userEmail, passwordHash = passwordHash, address = userAddress, contactNumber = userContactNo, emailHash = emailHash, educationalQualification =  educationQualification, clinicName = clinicName)
                    # try:
                    #     doctor.full_clean()
                    # except ValidationError:
                    #     message = "Please fill in the required fields (*)"
                    #     context = {
                    #                 "message": message
                    #             }

                    #     # Editing response headers so as to ignore cached versions of pages
                    #     response = render(request,"HealthCentre/registrationPortal.html",context)
                    #     return responseHeadersModifier(response)
                    doctor.save()
                    message = "Login Successful !"
                    global G_docName
                    DocObj = Doctor.objects.get(email = userEmail)
                    DocId = DocObj.pk
                    regDocName = name
                    G_docName = str(regDocName)
                    backgroundtastForQrCode()
            # Storing success message in the context variable
                    context = {
                        "userType" : userType,
                        "message": message
                    }

                    # Editing response headers so as to ignore cached versions of pages
                    response = render(request, "HealthCentre/registrationPortal.html",context)
                    # response = HttpResponseRedirect(reverse('catchqrcode'))
                    return responseHeadersModifier(response)
                except IntegrityError:
                    
                    message = "Your Name, Email or Contact Number already exists!"
                    context = {
                                "message": message
                            }

                    # Editing response headers so as to ignore cached versions of pages
                    response = render(request,"HealthCentre/registrationPortal.html",context)
                    return responseHeadersModifier(response)
                
            
                # if not wpIsConnected:
                #     request.session['wpStatus'] = "waiting for whatsapp QR code scan"
                # else:
                #     request.session['wpStatus'] = "whatsapp is registered"

            
            

            # Creating a patient object and saving insdie the database
            # patient = Patient(name = name,rollNumber = userRollNo, email = userEmail, passwordHash = passwordHash, address = userAddress, contactNumber = userContactNo, emailHash = emailHash )
            # patient.save()
            message = "Login Successful !"
            # Storing success message in the context variable
            context = {
                "userType" : userType,
                "message": message
            }

            # Editing response headers so as to ignore cached versions of pages
            # response = render(request, "HealthCentre/registrationPortal.html",context)
            response = HttpResponseRedirect(reverse('login'))
            return responseHeadersModifier(response)

        # If the passwords given don't match
        else:
            # Storing failure message in the context variable
            context = {
                "message":"Passwords do not match.Please register again."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request,"HealthCentre/registrationPortal.html",context)
            return responseHeadersModifier(response)

    # For any other method of request, sending back the registration page.
    else:

        # Editing response headers so as to ignore cached versions of pages
        response = render(request,"HealthCentre/registrationPortal.html")
        return responseHeadersModifier(response)


def responseHeadersModifier(response):
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

def doctors(request):
    """Function to send information about the doctors available to the user. """
    context = {
        "doctors": Doctor.objects.all()
    }
    response = render(request, "HealthCentre/contactus.html", context)
    return responseHeadersModifier(response)

def editDocDetails(request, pk):
    doctor = get_object_or_404(Doctor, id=pk)
    if request.method == "POST":
        doctor.name = request.POST.get('name', doctor.name)
        doctor.clinicName = request.POST.get('clinicName', doctor.clinicName)
        doctor.educationalQualification = request.POST.get('educationalQualification', doctor.educationalQualification)
        doctor.specialization = request.POST.get('specialization', doctor.specialization)
        doctor.address = request.POST.get('address', doctor.address)
        doctor.email = request.POST.get('email', doctor.email)
        doctor.contactNumber = request.POST.get('contactNumber', doctor.contactNumber)
        doctor.save()
        
        return redirect('doctors')
    
    doctorSpecific = Doctor.objects.filter(name=request.session.get('Name')).order_by('name')
    context = {
        "doctors": doctorSpecific
    }
    response = render(request, "HealthCentre/editDoctordetails.html", context)
    return responseHeadersModifier(response)

def yourPrescriptions(request):
    if request.method == "GET":
        doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
        records = doctor.doctorRecords.all()
        doctorSpecific = Prescription.objects.filter(prescribingDoctor = request.session['Name']).order_by('timestamp')
        context = {
                    "message" : "Successfully Logged In.",
                    "isAuthenticated" : True,
                    "user": records.order_by('-timestamp'),
                    "prescriptions" : doctorSpecific,
                    "prescMedicine" : Medicine.objects.all().order_by('id'),
                }
        response = render(request,"HealthCentre/prescriptionPortal.html", context)
        return responseHeadersModifier(response)
        
def login(request):
    """ Function for logging in the user. """

    # Calling session variables checker
    request = requestSessionInitializedChecker(request)
    # openWhatsapp.wp()
    
    # If the request method is post
    if request.method == "GET":
        try:

            # If the user is already logged in inside of his sessions, and is a doctor, then no authentication required
            if request.session['isLoggedIn'] and request.session['isDoctor']:
                # Accessing the doctor user and all his/her records
                doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
                records = doctor.doctorRecords.all()
                # Getting the count of the new prescriptions pending
                numberNewPendingPrescriptions = doctor.doctorRecords.aggregate(newnewPendingPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = False) ) ))['newnewPendingPrescriptions']
                # Storing the same inside the session variables
                request.session['numberNewPrescriptions'] = numberNewPendingPrescriptions
                doctorSpecific = Prescription.objects.filter(prescribingDoctor = request.session['Name']).order_by('timestamp')
                # request.session['qrcode'] = generateqr
                # Storing the required information inside the context variable
                context = {
                    "message" : "Successfully Logged In.",
                    "isAuthenticated" : True,
                    "user": records.order_by('-timestamp'),
                    "prescriptions" : doctorSpecific,
                    "prescMedicine" : Medicine.objects.all().order_by('id'),
                    
                }
                
                # Editing response headers so as to ignore cached versions of pages
                # response = render(request,"HealthCentre/prescriptionPortal.html", context)
                response = render(request,"HealthCentre/index.html", context)
                return responseHeadersModifier(response)
            
            # If the user is already logged in inside of his sessions, and is a patient, then no authentication required
            elif request.session['isLoggedIn'] and (not request.session['isDoctor']):

                # Accessing the patient user and all his/her records
                patient = Patient.objects.get(emailHash = request.session['userEmail'])
                records = patient.patientRecords.all()

                # Getting the count of the new prescriptions pending
                numberNewPrescriptions = patient.patientRecords.aggregate(newCompletedPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = True) ) ) )['newCompletedPrescriptions']

                # Storing the same inside the session variables
                request.session['numberNewPrescriptions'] = numberNewPrescriptions

                # Updating the completed records
                for record in records:
                    if record.isCompleted:
                        record.isNew = False
                        record.save()

                # Storing the required information inside the context variable
                context = {
                    "message" : "Successfully Logged In.",
                    "isAuthenticated" : True,
                    "user": records.order_by('-timestamp')
                    }

                # Editing response headers so as to ignore cached versions of pages
                # response = render(request,"HealthCentre/prescriptionPortal.html")
                response = render(request,"HealthCentre/userPatientProfilePortal.html", context)
                return responseHeadersModifier(response)

            else:
                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/loginPortal.html")
                return responseHeadersModifier(response)

        # If any error occurs, sending back a new blank page
        except:

            # Editing response headers so as to ignore cached versions of pages
            response = render(request,"HealthCentre/loginPortal.html")
            return responseHeadersModifier(response)
        
    # If the request method is post
    elif request.method == "POST":

        # Extracting the user information from the post request
        userName = request.POST["useremail"]
        userPassword = request.POST["userpassword"]

        # If such a patient exists
        try:
            patient = Patient.objects.get(email = userName)

            # Storing required session information
            request.session['isDoctor'] = False

        # Otherwise trying if a doctor exists
        except Patient.DoesNotExist:
            try:
                doctor = Doctor.objects.get(email = userName)

                # Storing required session information
                request.session['isDoctor'] = True     

            # If no such doctor or patient exists
            except Doctor.DoesNotExist:
                adminUsername = "admin@123"
                adminPassword = "12345"
                if userName == adminUsername and userPassword == adminPassword:
                    
                    # return render(request, "HealthCentre/adminPortal.html")
                    return HttpResponseRedirect(reverse("admin"))
                # Storing message inside context variable
                context = {
                    "message":"User does not exist.Please register first."
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/loginPortal.html", context)
                return responseHeadersModifier(response)

        # Getting the hash of user inputted password
        passwordHash = passwordHasher(userPassword)

        # If the logged in user is a doctor
        if request.session['isDoctor']:
            
            # Accessing all records of doctor
            records = doctor.doctorRecords.all()

            # Getting the count of new prescriptions
            numberNewPendingPrescriptions = doctor.doctorRecords.aggregate(newPendingPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = False) ) ))['newPendingPrescriptions']

            # Storing the same inside request variable
            request.session['numberNewPrescriptions'] = numberNewPendingPrescriptions

            # If the inputted hash and the original user password hash match
            if passwordHash == doctor.passwordHash:

                # Storing required information in session variable of request
                request.session['isLoggedIn'] = True
                request.session['userEmail'] = doctor.emailHash
                IDofDoc = doctor.pk
                docId = Doctor.objects.get(id = IDofDoc)
                request.session['Name'] = docId.name
                global G_docName
                docNameCommon = request.session['Name']
                G_docName = str(docNameCommon)
                
                # Redirecting to avoid form resubmission
                # Redirecting to home page
                # Editing response headers so as to ignore cached versions of pages
                # response = render(request,"HealthCentre/userDoctorProfilePortal.html")
                # response = HttpResponseRedirect(reverse('onlineprescription'))
                # response = render(request,"HealthCentre/prescriptionPortal.html")
                response = HttpResponseRedirect(reverse('index'))
                return responseHeadersModifier(response)

            # Else if the password inputted is worng and doesn't match
            else:

                # Storing message inside context variable
                context = {
                    "message":"Invalid Credentials.Please Try Again."
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/loginPortal.html", context)
                return responseHeadersModifier(response)

        # Otherwise if the user is a patient
        else:

            # Accessing all records of patient
            records = patient.patientRecords.all()

            # Getting the count of new prescriptions
            numberNewPrescriptions = patient.patientRecords.aggregate(newCompletedPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = True) ) ))['newCompletedPrescriptions']

            # Storing the same inside request variable
            request.session['numberNewPrescriptions'] = numberNewPrescriptions

            # Updating the completed records
            for record in records:
                if record.isCompleted :
                    record.isNew = False
                    record.save()

            # If the inputted hash and the original user password hash match
            if passwordHash == patient.passwordHash:

                # Storing required information in session variable of request
                request.session['isLoggedIn'] = True
                request.session['userEmail'] = patient.emailHash
                request.session['Name'] = patient.name
                request.session['isDoctor'] = False

                # Redirecting to avoid form resubmission
                # Redirecting to home page
                # Editing response headers so as to ignore cached versions of pages
                # response = render(request,"HealthCentre/userPatientProfilePortal.html")
                # response = render(request, "HealthCentre/prescriptionportal.html")
                # response = HttpResponseRedirect(reverse('onlineprescription'))
                response = HttpResponseRedirect(reverse('index'))
                return responseHeadersModifier(response)

            # Else if the password inputted is worng and doesn't match
            else:

                # Storing message inside context variable
                context = {
                    "message":"Invalid Credentials.Please Try Again."
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/loginPortal.html", context)
                return responseHeadersModifier(response)
    # For any other method of access, returning a new blank login page
    else:
        response = render(request,"HealthCentre/loginPortal.html")
        return responseHeadersModifier(response)
    
def forgot_password(request):
    
    if request.method == 'POST':
        form = ForgotPasswordForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            docMail = Doctor.objects.filter(email=email)
                    # Handle email sending failure
            if not docMail.exists():        
                form.add_error(None, 'This mail is not registered')
                return render(request, 'HealthCentre/forgot_password.html', {'message': 'This email is not registered'})

                    # print(f"Error sending email: {e}")
            for doctremail in docMail:
                docQuerymail = doctremail.email
            # docMail = Doctor.objects.filter(email=email).first()
            if email == docQuerymail:
                # Generate OTP
                otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
                
                try:
                    # Send OTP to user's email
                    send_mail(
                        'Password Reset OTP',
                        f'Your OTP for password reset is: {otp}',
                        settings.EMAIL_HOST_USER,
                        [email],
                        fail_silently=False,
                    )
                    # Store the OTP in session for verification later
                    request.session['otp'] = otp
                    request.session['email'] = email
                    return redirect('verify_otp')
                except Exception as e:
                    # Handle email sending failure
                    form.add_error(None, 'There was an error sending the email. Please try again.')
                    print(f"Error sending email: {e}")
            else:
                form.add_error('email', 'Please enter the valid email id')
    else:
        form = ForgotPasswordForm()
    
    return render(request, 'HealthCentre/forgot_password.html', {'form': form})
# def forgot_password(request,pk):
#     docEmail = Doctor.objects.get(id=pk)
#     docMailid= docEmail.email
#     if request.method == 'POST':
#         form = ForgotPasswordForm(request.POST)
#         if form.is_valid():
#             email = form.cleaned_data['email']

#             if email == docMailid:
#             # Generate OTP
#                 otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
            
#                 try:
#                     # Send OTP to user's email
#                     send_mail(
#                         'Password Reset OTP',
#                         f'Your OTP for password reset is: {otp}',
#                         settings.EMAIL_HOST_USER,
#                         [email],
#                         fail_silently=False,
#                     )
#                     # Store the OTP in session for verification later
#                     request.session['otp'] = otp
#                     request.session['email'] = email
#                     return redirect('verify_otp')
#                 except Exception as e:
#                 # Handle email sending failure
#                     form.add_error(None, 'There was an error sending the email. Please try again.')
#                     print(f"Error sending email: {e}")
#                 else:
#                     print("Please enter the valid email id")    
        
#     else:
#         form = ForgotPasswordForm()
#     return render(request, 'HealthCentre/forgot_password.html', {'form': form})

def verify_otp(request):
    if request.method == 'POST':
        form = VerifyOTPForm(request.POST)
        if form.is_valid():
            user_otp = form.cleaned_data['otp']
            stored_otp = request.session.get('otp')
            email = request.session.get('email')
            if user_otp == stored_otp:
                # If OTP matches, proceed with password reset
                # Clear OTP and email from session
                del request.session['otp']
                del request.session['email']
                return redirect('reset_password', email=email)
            else:
                # If OTP doesn't match, show error message
                return render(request, 'HealthCentre/verify_otp.html', {'form': form, 'error': 'Invalid OTP. Please try again.'})
    else:
        form = VerifyOTPForm()
    # return render(request, 'HealthCentre/verify_otp.html', {'form': form})
    return render(request, 'HealthCentre/forgot_password.html', {'form': form})


def reset_password(request, email):
    if request.method == 'POST':
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            passHash = passwordHasher(new_password)
            # Reset the password for the user with the provided email
            try:
                user = Doctor.objects.get(email=email)
                user.passwordHash = passHash
                user.save()
                return HttpResponseRedirect(reverse("index"))
                # return render(request, 'HealthCentre/password_reset_success.html')
            except Doctor.DoesNotExist:
                pass  # Handle case where user doesn't exist
    else:
        form = ResetPasswordForm()
    return render(request, 'HealthCentre/reset_password.html', {'form': form})


def admin(request):
    
    return render(request, "HealthCentre/adminPortal.html")
    

def emergency(request):
    """ Funtion for emergency situations, for requesting an ambulance."""

    # If the request method is get
    if request.method == "GET":

        # Editing response headers so as to ignore cached versions of pages
        response = render(request,"HealthCentre/emergencyPortal.html")
        return responseHeadersModifier(response)

    # If the request method is post and the user is submitting information
    elif request.method == "POST":

        # Extracting the emergency location from the post request
        emergencyLocation = request.POST['emergencyLocation']

        # Giving emergency message to server, can also be connected to IOT devices for alarms
        # If the emergency location text is not an empty string
        if emergencyLocation != "":

            # Printing information and notifying inside of server terminal
            print("------------------------------------------------------------------------")
            print("\n\nEMERGENCY !! AMBULANCE REQUIRED AT " + emergencyLocation + " !!\n\n")
            print("------------------------------------------------------------------------")

            # Storing information inside of context variable
            context = {
                "message" : "Ambulance reaching " + emergencyLocation + " in 2 minutes."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/emergencyPortal.html", context)
            return responseHeadersModifier(response)

        # Else if the emergency location is an empty string
        else:

            # Storing message inside context variable
            context = {
                "message" : "No location entered.Invalid input."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/emergencyPortal.html", context)
            return responseHeadersModifier(response)

    # For any other method of access, returning a new blank emergency page
    else:

        # Editing response headers so as to ignore cached versions of pages
        response = render(request,"HealthCentre/emergencyPortal.html")
        return responseHeadersModifier(response)

def logout(request):
    """Function to log out the user."""
    # Erasing all the information of the session variables if user is logged out
    request.session['isDoctor'] = ""
    request.session['isLoggedIn'] = False
    request.session['userEmail'] = ""
    request.session['Name'] = ""
    request.session['numberNewPrescriptions'] = ""
    request.session['writeNewPrescription'] = False
    request.session['createNewAppointment'] = False

    # Redirecting to avoid form resubmission
    # Redirecting to home page
    # Editing response headers so as to ignore cached versions of pages
    response = HttpResponseRedirect(reverse('login'))
    return responseHeadersModifier(response)

def contactus(request):
    """Function to display contact information."""

    # Editing response headers so as to ignore cached versions of pages
    response = render(request, "HealthCentre/contactus.html")
    return responseHeadersModifier(response)


def doctorappointmentsfalse(request):
    if request.method == 'GET':
        # request.session['goToAppointmentsPage'] = True
        request.session['createNewAppointment'] = True
        if request.session['isLoggedIn'] and request.session['isDoctor'] and request.session['createNewAppointment']:
            # Accessing the doctor user and all his/her records
            # request.session['CreatenewAppointment'] = False
            doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
            records = doctor.doctorRecords.all()
            doctorSpecific = Appointment.objects.filter(appointmentdoctor = request.session['Name']).order_by('-date')
            # Getting the count of the new prescriptions pending
            numberNewPendingPrescriptions = doctor.doctorRecords.aggregate(newnewPendingPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = False) ) ))['newnewPendingPrescriptions']

            # Storing the same inside the session variables
            request.session['numberNewPrescriptions'] = numberNewPendingPrescriptions

            # Storing the required information inside the context variable
            context = {
                "message" : "Successfully Logged In.",
                "isAuthenticated" : True,
                "user": records.order_by('-timestamp'),
                # "Appointments" : Appointment.objects.all().order_by('-date')
                "Appointments" : doctorSpecific
            }
            response = render(request,"HealthCentre/appointmentsPortal.html", context)
            return responseHeadersModifier(response)
        
def doctorappointments(request):
    if request.method == 'GET':
        request.session['goToAppointmentsPage'] = True
        request.session['appointmentEdit'] = False
        # request.session['createNewAppointment'] = True
        form = AppointmentForm()
        model = AppointmentForm()
        # form = AppointmentForm(request.POST or None)
        hour = range(00, 24)
        minute = range(00, 60)
        date = range(1, 32)
        month = range(1, 13)
        year = range(int(datetime.now().year), 2099)
        doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
        context = {'form': form, 
                    'model': model,
                    'hours': hour,
                    'dates' : date,
                    'months' : month,
                    'years' : year,
                    'minutes': minute,
                    "patients" : doctorSpecific, #Patient.objects.all().order_by('id'),
                    # "prescPatients" : Appointment.objects.all().order_by('id')
                    }
        response = render(request, 'HealthCentre/NewAppointment.html', context)
        return responseHeadersModifier(response)
    if request.method == 'POST':
        currentDateTime = datetime.now()
        currentHourObj = datetime.strftime(currentDateTime, "%H")
        currentMinuteObj = datetime.strftime(currentDateTime, "%M")
        currentDateObj = datetime.strftime(currentDateTime, "%d")
        currentMonthObj = datetime.strftime(currentDateTime, "%m")
        currentYearObj = datetime.strftime(currentDateTime, "%Y")
        datePick = request.POST['datePick']
        timePick = request.POST['timePick']

        
        if request.session['goToAppointmentsPage']:
            if request.POST['selectedPatient'] == "":
                appointmentPatient = request.POST['PatientNameForAppointment']
                # patient = Patient.objects.create(name=prescpatient)
            else:
                appointmentPatient = request.POST['selectedPatient']
                # prescpatient = request.POST['selectedPatient']
                patient_id = request.POST['selectedPatient'] 
                docName = request.session['Name']
                patient = Patient.objects.get(name=patient_id, doctorname = docName)
            # appointmentTime = datetime.(timePick) #request.POST['EnterTimeHour'].zfill(2) + request.POST['EnterTimeMinute'].zfill(2)
            datetimeObject = datetime.strptime(timePick, "%H:%M") #, "%H%M"
            datetimeObject = datetimeObject.time()
            combinetime = datetime.combine(datetime.today(), datetimeObject)
            AppointmentTime1 = datetime.strftime(combinetime, '%I:%M %p')
        
            currentDateTimeObj = currentDateObj + currentMonthObj + currentYearObj
            currentDateTimeObj = datetime.strptime(currentDateTimeObj, "%d%m%Y")
            currentDateTimeObj = currentDateTimeObj.date()
            
            currentTimeObj = currentHourObj + currentMinuteObj
            currentTimeObj = datetime.strptime(currentTimeObj, "%H%M")
            currentTimeObj1 = currentTimeObj.time()
            currentTimeObjPlusThreeHours = currentTimeObj + timedelta(hours= 3)
            currentTimeObjPlusThreeHours1 = currentTimeObjPlusThreeHours.time()
            hourObject = datetime.strftime(currentTimeObjPlusThreeHours, "%H")
            minuteObject = datetime.strftime(currentTimeObjPlusThreeHours, "%M")
            strTimeObject = hourObject + minuteObject
            datetimeObjectplusThree = datetime.strptime(strTimeObject, "%H%M")
            datetimeObjectplusThree = datetimeObjectplusThree.time()
            
            # appointmentDate = request.POST['EnterDate'] + request.POST['EnterDateMonth'] + request.POST['EnterYear']
            dateobject = datetime.strptime(datePick, "%Y-%m-%d") #, "%d%m%Y")
            dateobject = dateobject.date()
            combinedate = datetime.combine(dateobject, datetime.now().time())
            AppointmentDate1 = datetime.strftime(combinedate, "%b. %d, %Y")
            currentDateObject = currentDateObj + currentMonthObj + currentYearObj 
            currentDateObject = datetime.strptime(currentDateObject, "%d%m%Y")
            
            appointmentNotes = request.POST['AppointmentDescription']
            appointmentDoctor = request.session['Name']
            appointmentSubject = "subject"
            doctor_id = request.session['Name']
            doctorid = Doctor.objects.get(name=doctor_id)
                # existingDate = Appointment.objects.filter(date = dateobject)
            try:
                existingTime = Appointment.objects.get(Q(time = datetimeObject) & Q(date = dateobject) & Q(appointmentdoctor = request.session['Name']))
            except Appointment.DoesNotExist:
                existingTime = None
                pass
                    
            # for extime in existingTime:
            #     if existingTime == None:
            #         pass
            if existingTime != None:
                existingAppointmentTime = existingTime.time # extime.time
                # existingAppointmentTime = existingAppointmentTime.time()
                existingAppointmentDate = existingTime.date #  extime.date
                # existingAppointmentDate = existingAppointmentDate.time()
                if existingAppointmentTime == datetimeObject and existingAppointmentDate == dateobject:
                    existingAppointmentStatus = "You already have another appointment at this time! Please set another time"
                    doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
                    records = doctor.doctorRecords.all()
                    context = {
                            "message" : "Successfully Logged In.",
                            "isAuthenticated" : True,
                            "user": records.order_by('-timestamp'),
                            "Appointments" : Appointment.objects.all().order_by('-date'),
                            "existingAppointmentStatus" : existingAppointmentStatus
                            }
            
                    response = render(request,'HealthCentre/NewAppointment.html', context)
                    return responseHeadersModifier(response)
            else :
                appointment = Appointment(time = timePick, date = datePick, subject = appointmentSubject, notes = appointmentNotes,
                                            appointmentpatient = appointmentPatient, appointmentdoctor = appointmentDoctor, doctorPres = doctorid,
                                            patientPres = patient)
                appointment.save()
                doctorDetail = Doctor.objects.get(name = doctor_id)
                doctorNumber = doctorDetail.contactNumber
                patientDetail = Patient.objects.get(name = patient_id, doctorname = doctor_id)
                patientNumber = patientDetail.contactNumber
                if (datetimeObject < datetimeObjectplusThree) and (datetimeObject > currentTimeObj1) and (currentDateTimeObj == dateobject):
                    whatsappApi(patient_id, doctor_id,  patientNumber, AppointmentTime1, AppointmentDate1, doctorDetail.clinicName) # datetimeObject, dateobject)
                    whatsappApiDoc(doctor_id, doctorNumber, AppointmentTime1, AppointmentDate1, patient_id, patientNumber) # datetimeObject, dateobject)
                    # time.sleep(60)
                    
                doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
                records = doctor.doctorRecords.all()
                context = {
                "message" : "Successfully Logged In.",
                "isAuthenticated" : True,
                "user": records.order_by('-timestamp'),
                "Appointments" : Appointment.objects.all().order_by('-date')
            }
            # whatsappNotification()
                response = HttpResponseRedirect(reverse('doctorappointmentsfalse'))
                return responseHeadersModifier(response)

def editAppointments(request, pk):
    request.session['appointmentEdit'] = True
    # record = get_object_or_404(Appointment, pk = record_id)
    appointment = Appointment.objects.get(id=pk)
    form = AppointmentForm(instance=appointment)
    if request.method == "POST":
        # form = AppointmentForm(request.POST, instance=appointment)
        # if form.is_valid():
            # form.save()
        
        # records = doctor.doctorRecords.all()
        appointObject = Appointment.objects.get(id=pk)
        if request.POST['selectedPatient'] == "":
            appointObject.appointmentpatient = request.POST['PatientNameForAppointment']
            # patient = Patient.objects.create(name=prescpatient)
        else:
            appointObject.appointmentpatient = request.POST['selectedPatient']
            # prescpatient = request.POST['selectedPatient']
            # patient_id = request.POST['selectedPatient'] 
            # patient = Patient.objects.get(name=patient_id)
        
        appointmentTime = request.POST['EnterTimeHour'].zfill(2) + request.POST['EnterTimeMinute'].zfill(2)
        appointmentTime = datetime.strptime(appointmentTime, "%H%M")
        appointObject.time = appointmentTime.time()
        combinetime = datetime.combine(datetime.today(), appointObject.time)
        AppointmentTime1 = datetime.strftime(combinetime, '%I:%M %p')
        appointmentDate = request.POST['EnterDate'].zfill(2) + request.POST['EnterDateMonth'].zfill(2) + request.POST['EnterYear'].zfill(2)
        appointmentDate = datetime.strptime(appointmentDate, "%d%m%Y")
        appointObject.date = appointmentDate
        combinedate = datetime.combine(appointObject.date, datetime.now().time())
        AppointmentDate1 = datetime.strftime(combinedate, "%b. %d, %Y")
        appointObject.notes = request.POST['AppointmentDescription']
        appointObject.appointmentdoctor = request.session['Name']
        appointObject.subject = "subject"
        appointObject.save()
        patientDetails = Patient.objects.get(name=appointObject.appointmentpatient)
        patientNumber = patientDetails.contactNumber
        doctor = Doctor.objects.get(name = request.session['Name'])
        # thread = threading.Thread(target = whatsappApiEdit, args = (appointObject.appointmentpatient, appointObject.appointmentdoctor, patientNumber, AppointmentTime1, AppointmentDate1, doctor.clinicName)) # appointObject.time, appointObject.date)
        # thread.start()
        # appointment = Appointment(time = datetimeObject, date = dateobject, subject = appointmentSubject, notes = appointmentNotes,
        #                             appointmentpatient = appointmentPatient, appointmentdoctor = appointmentDoctor)
        # appointment.save()
        whatsappApiEdit(appointObject.appointmentpatient, appointObject.appointmentdoctor, patientNumber, AppointmentTime1, AppointmentDate1, doctor.clinicName) # appointObject.time, appointObject.date)
        doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
        records = doctor.doctorRecords.all()
        context = {
        "message" : "Successfully Logged In.",
        "isAuthenticated" : True,
        "user": records.order_by('-timestamp'),
        "Appointments" : Appointment.objects.all().order_by('-date')
    }
        # response = HttpResponseRedirect(request, 'HealthCentre/appointmentsPortal.html', context)
        response = HttpResponseRedirect(reverse('doctorappointmentsfalse'))
        return responseHeadersModifier(response)
        
    # hour = Appointment.objects.get()
    # hour = Appointment.objects.get(time=time)
    hour = range(00, 24)
    minute = range(00, 60)
    date = range(1, 32)
    month = range(1, 13)
    year = range(int(datetime.now().year), 2099)
    appointmentObject = Appointment.objects.get(id=pk)
    patient = appointmentObject.appointmentpatient
    datobject = appointmentObject.date
    dateobject = datetime.strftime(datobject, "%d")
    # monthobject = appointmentObject.date
    monthobject = datetime.strftime(datobject, "%m")
    # yearobject = appointmentObject.date
    yearobject = datetime.strftime(datobject, "%Y")
    timeobject = datetime.combine(appointmentObject.date, appointmentObject.time) 
    minuteobject = datetime.strftime(timeobject, "%M")
    hourobject = datetime.strftime(timeobject, "%H")
    appointmentnotes = appointmentObject.notes
    context = {
        
            'hours': hour,
            'editDate' : dateobject,
            'dates' : date,
            'editMonth' : monthobject,
            'months' : month,
            'editYear' : yearobject,
            'years' : year,
            'editMinute' : minuteobject,
            'minutes': minute,
            'editHour' : hourobject,
            "patients" : patient,
            "editNotes" : appointmentnotes,
            "pats" : Patient.objects.all().order_by('id'),
            "prescPatients" : Appointment.objects.all().order_by('id'),
            'form' : form
    }
    # appointmentObject.date = datetime.strptime((request.POST.get('EnterDate') + request.POST.get('EnterDateMonth') + request.POST.get('EnterYear')), "%m%d%Y") 
    # appointmentObject.date = datetime.strptime(appointmentObject.date, "%m%d%Y")
    # appointmentObject.save()
    # return render(request,'HealthCentre/NewAppointment.html', context)
    response = render(request,'HealthCentre/NewAppointment.html', context)
    return responseHeadersModifier(response)

def deleteappointment(request, pk):
    request.session['deleteAppointment'] = True
    delappointmentobj = Appointment.objects.get(id=pk)
    delappointmentobj.delete()
    doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
    records = doctor.doctorRecords.all()
    context = {
    "message" : "Successfully Logged In.",
    "isAuthenticated" : True,
    "user": records.order_by('-timestamp'),
    "Appointments" : Appointment.objects.all().order_by('-date')
        }
    response = response = HttpResponseRedirect(reverse('doctorappointmentsfalse'))
    return responseHeadersModifier(response)

rowCount = 0
additionalRowData ={}
selectedMedicineID = []
selectedSessionID = []
@csrf_exempt    
def createNewMedicine(request):
    if request.method == "POST":
        print(request.POST)
        newMedicine = request.POST.get("NewmedicineName")
        befAft = request.POST.get("befAftFood")
        
        medicine = Medicine(medicinename = newMedicine, beforeafter = befAft, morning = "0", afternoon = "0", night = "0")
        medicine.save()
        data ={
            "newMedicine" :newMedicine,
            "befAft": befAft
        }
    # response = HttpResponseRedirect(reverse('doctorprofile'))
    return JsonResponse(data)

def addingSessionData(request, SelectedSessionValue):
    try:
        SelectedSession = timeofday.objects.get(timeoftheday = SelectedSessionValue)
        request.session['selectedSession'] = SelectedSessionValue
        SelectedMorning = SelectedSessionValue[0]
        SelectedAfternoon = SelectedSessionValue[3]
        SelectedNight = SelectedSessionValue[6]
        sessionID = SelectedSession.pk
        selectedSessionID.append(sessionID)
        request.session['sessionIsSelected'] = True
        request.session['clickedOnAddRow'] = True
        if request.session['medicineIsSelected'] == True:
            clickedOnAddRow = True
            medicineIsSelected = True
            sessionIsSelected = True
        data = {
                    "MedMorn" : SelectedMorning,
                    "medAft"   : SelectedAfternoon,
                    "medNight" : SelectedNight,
                    'clickedOnAddRow' : clickedOnAddRow,
                    'medicineIsSelected' : medicineIsSelected,
                    'sessionIsSelected' : sessionIsSelected
                }
        request.session['additionalSession_data'] = data
        global additionalRowData
        addedData = request.session.get('additionalMedicine_data', {})
        addedData.update(data)
        for i in range(int(rowCount)):
            addingData = {
                f'MedBefAft{i+1}' : addedData.get('MedBefAft'),
                f'SelectedMed{i+1}': addedData.get('SelectedMed'),
                f'MedMorn{i+1}': addedData.get('MedMorn'),
                f'medAft{i+1}': addedData.get('medAft'),
                f'medNight{i+1}' : addedData.get('medNight')
            }
        additionalRowData.update(addingData)
        
        return JsonResponse(data)
            # response = render(request, "HealthCentre/NewPrescription.html", data)
            # return responseHeadersModifier(response)
    except timeofday.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)

def addingMedicineData(request, selectedMedicineValue):
        global rowCount
        try:
            SelectedMedicine = Medicine.objects.get(medicinename = selectedMedicineValue)
            
            SelectedBeforeAfter = SelectedMedicine.beforeafter
            # SelectedMorning = SelectedMedicine.Morning
            # SelectedAfternoon = SelectedMedicine.Afternoon
            # SelectedNight = SelectedMedicine.Night
            
            medicineID = SelectedMedicine.pk
            
            selectedMedicineID.append(medicineID)
            request.session['medicineIsSelected'] = True
            request.session['clickedOnAddRow'] = True
            # prescription = Prescription.objects.get(id= getPrescriptionID)
            # prescription.medicine.add(medicineID)
            data = {
                "MedBefAft" : SelectedBeforeAfter,
                "SelectedMed" : selectedMedicineValue,
                'clickedOnAddRow' : True,
                'medicineIsSelected' : True,
                
                # "MedMorn" : SelectedMorning,
                # "medAft"   : SelectedAfternoon,
                # "medNight" : SelectedNight,
                # "patients" : Patient.objects.all().order_by('id'),
                # "prescPatients" : Prescription.objects.all().order_by('id'),
                # "prescMedicines" : Medicine.objects.all().order_by('MedicineName')
            }
            request.session['additionalMedicine_data'] = data
            
            return JsonResponse(data)
            # response = render(request, "HealthCentre/NewPrescription.html", data)
            # return responseHeadersModifier(response)
        except Medicine.DoesNotExist:
                return JsonResponse({'error': 'Medicine not found'}, status=404)

# dummyBoolean = False
global dummyBoolean
dummyBoolean = False

def dummy(request):
    if request.method == 'GET':
        global dummyBoolean
        dummyBoolean = False
    if request.method == 'POST':
        dummyBoolean = True
    return HttpResponseRedirect(reverse("doctorprofile"))
    

def doctorprofile(request):
    #  selectedMedicineValue = ""
     if request.method == 'GET':
        request.session['writeNewPrescription'] = True
        
        if request.GET.get('noofdays') == None and request.GET.get('SelectedMed') == None and request.GET.get('SelectedPat') == None and request.GET.get('SelectedSess') == None:
            docn = request.session['Name']
            doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
            clickedOnAddRow = True
            medicineIsSelected = True
            sessionIsSelected = True
            # doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
            try:
                doctor = doctorlogo.objects.get(docname=request.session['Name'])
                docLogo = doctor.logo
                stringPath = os.path.abspath(str(docLogo))
                absPath = docLogo.path.split(os.path.sep)
                absPathAsString = str(stringPath)
                pathIndex = absPath.index("static")
                newPath = os.path.join(*absPath[pathIndex:])
                imagedata = {
                    "doclogo" : absPathAsString
                }
                request.session['imagePath'] = imagedata
            except doctorlogo.DoesNotExist:
                newPath = ''
            context = {
                    "patients" : doctorSpecific, #Patient.objects.all().order_by('id'),
                    # "prescPatients" : Prescription.objects.all().order_by('id'),
                    "prescMedicines" : Medicine.objects.all().order_by('medicinename'),
                    "prescTimeOfDay" : timeofday.objects.all().order_by('id'),
                    'clickedOnAddRow' : clickedOnAddRow,
                    'medicineIsSelected' : medicineIsSelected,
                    'sessionIsSelected' : sessionIsSelected,
                    "doclogo" : newPath,
                    }
            response = render(request, "HealthCentre/NewPrescription.html", context)
            return responseHeadersModifier(response)
        if request.GET.get('SelectedPat') != None and request.method == 'GET':
            PatientName = request.GET.get('SelectedPat', None)
            try: 
                selectedPatient = Patient.objects.get(name = PatientName)
                data = {
                    "patientName": selectedPatient.name,
                    "patientSex" : selectedPatient.passwordHash,
                    "patientAge" : selectedPatient.rollNumber
                }
                request.session['pdf_data'] = data
                return JsonResponse(data)
            except Patient.DoesNotExist:
                return JsonResponse({'error': 'Patient not found'}, status=404)
            
        if request.GET.get('SelectedSess') != None and request.method == 'GET':
            SesssionTime = request.GET.get('SelectedSess', None)
            
            try:
                SelectedSession = timeofday.objects.get(timeoftheday = SesssionTime)
                sessionID = SelectedSession.pk
                selectedSessionID.append(sessionID)
                SelectedMorning = SesssionTime[0]
                SelectedAfternoon = SesssionTime[3]
                SelectedNight = SesssionTime[6]
                data = {
                        "MedMorn" : SelectedMorning,
                        "medAft"   : SelectedAfternoon,
                        "medNight" : SelectedNight,
                    }
                request.session['pdf_sess_data'] = data
                # global renderedHTML
                # # renderedHTMLbuffer = renderedHTML
                # renderedHTML = render_to_string('NewPrescription.html', data)
                return JsonResponse(data)
            except timeofday.DoesNotExist:
                return JsonResponse({'error': 'Session not found'}, status=404)
        
        if request.GET.get('SelectedMed') != None and request.method == 'GET':
    
            MedicineName = request.GET.get('SelectedMed', None)
            
            try:
                # SelectedMedicine = Medicine.objects.get(MedicineName = selectedMedicineValue)
                SelectedMedicine = Medicine.objects.get(medicinename = MedicineName)
                medicineID = SelectedMedicine.pk
                selectedMedicineID.append(medicineID)
                SelectedBeforeAfter = SelectedMedicine.beforeafter
                data = {
                    "MedBefAft" : SelectedBeforeAfter,
                    "SelectedMed" : SelectedMedicine.medicinename,
                    # "MedMorn" : SelectedMorning,
                    # "medAft"   : SelectedAfternoon,
                    # "medNight" : SelectedNight,
                }
                request.session['pdf_med_data'] = data
                # global renderedHTML
                # # renderedHTMLbuffer = renderedHTML
                # renderedHTML = render_to_string('NewPrescription.html', data)
                return JsonResponse(data)
            except Medicine.DoesNotExist:
                return JsonResponse({'error': 'Medicine not found'}, status=404)
            
        if request.GET.get('noofdays') != None and request.method == 'GET':
            noofdays = request.GET.get('noofdays', None)
            try:
                data = {
                    "noofdays" : noofdays
                }
                request.session['pdf_days_data'] = data
                return JsonResponse(data)
            except:
                return JsonResponse({'error': 'none value returned'}, status=404)
     if request.method == 'POST':
        
        if request.session['writeNewPrescription']:
            if request.POST['selectedPatient'] == "":
                prescpatient = request.POST['PatientName']
                patient = Patient.objects.create(name=prescpatient)
            else:
                prescpatient = request.POST['selectedPatient']
                patient_id = request.POST['selectedPatient']
                patient = Patient.objects.get(name=patient_id)

            symptoms = "dummy"#request.POST["symptoms"]
            if request.session['isLoggedIn'] and request.session['isDoctor']:
                prescdoctor = request.session['Name']
                # doctor = Doctor.objects.get(id=1)
                doctor_id = request.session['Name']
                doctor = Doctor.objects.get(name=doctor_id)
                medicine = request.POST['SelectedMedicine']
                MedicineObject = Medicine.objects.get(medicinename = medicine)
                
                selectedMedicines = Medicine.objects.filter(id__in = selectedMedicineID)
                selectedSessions = timeofday.objects.filter(id__in = selectedSessionID)
                # MedName = MedicineObject.MedicineName
                NoOfDays = request.POST['noOfDays']
                # patient_id = request.POST['selectedPatient'] 
                patientObj = Patient.objects.get(name=prescpatient)
                prescriptiontext = "dummy"#request.POST['prescription']
                prescription = Prescription(prescribingDoctor = prescdoctor, prescribingPatient = prescpatient ,doctor = doctor, patient= patient, symptoms = symptoms, prescriptionText = prescriptiontext, NoOfDays = NoOfDays) #medicine = medicine,
                prescription.save()
                global getPrescriptionID 
                getPrescriptionID = prescription.pk
                prescription.medicine.set(selectedMedicines)
                prescription.MornAftNight.set(selectedSessions)
                wpnumber = patientObj.contactNumber
                doctorSpecific = Prescription.objects.filter(prescribingDoctor = request.session['Name']).order_by('timestamp')
                docName = request.session['Name']
                Todaysdate = date.today()
                global dummyBoolean
                if dummyBoolean == True:
                    sendPdfinWhatsapp(wpnumber, docName, prescpatient, Todaysdate)
                clickedOnAddRow = True
                medicineIsSelected = True
                sessionIsSelected = True
            context = {
                    "prescriptions" : doctorSpecific,
                    'clickedOnAddRow' : clickedOnAddRow,
                    'medicineIsSelected' : medicineIsSelected,
                    'sessionIsSelected' : sessionIsSelected
                }
        response = render(request, "HealthCentre/prescriptionportal.html", context)
        return responseHeadersModifier(response)

def uploadImage(request):
    
    if request.method == 'GET':
        response = HttpResponseRedirect(reverse('doctorprofile'))
        return responseHeadersModifier(response)
    if request.method == 'POST':
        nameofdoctor = request.session['Name']
        image_file = request.FILES.get('LogoImage')
        try:
            logoobj = doctorlogo.objects.get(docname = nameofdoctor)
            if logoobj.logo != image_file:
                logoobj.logo = image_file
                logoobj.save()
                logoOfDoc = logoobj.logo
            elif logoobj.logo == image_file:
                logoobj.logo.delete()
                logoobj.logo = image_file
                logoobj.save()
                logoOfDoc = logoobj.logo
            context ={
                    "doclogo" :logoOfDoc,
                        
                    }
            
            response = HttpResponseRedirect(reverse('doctorprofile'))
            return responseHeadersModifier(response)
        except doctorlogo.DoesNotExist:
            doctorObj = Doctor.objects.get(name = nameofdoctor)
            docID = doctorObj.pk
            doctorlogoupload = doctorlogo(logo= image_file, docname = nameofdoctor, doctorid= doctorObj )
            doctorlogoupload.save()
            logoOfDoc = doctorlogoupload.logo

            context ={
                        "doclogo" :logoOfDoc,
                            
                        }
            response = render(request, 'HealthCentre/NewPrescription.html',context)
            return responseHeadersModifier(response)

def createTimeline(request):
    if request.method == 'GET':

        if request.GET.get('SelectedPat') == None:
                doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
                context = {
                        "patients" : doctorSpecific,
                        "patientName" : ""
                        }
                response = render(request, "HealthCentre/timeline.html", context)
                return responseHeadersModifier(response)
    if request.method == 'POST':
            # PatientName = request.GET.get('SelectedPat', None)
            PatientName = request.POST['selectedPatient']
            try: 
                selectedPatient = Patient.objects.get(name = PatientName, doctorname = request.session['Name'])
                selectedPatientID = selectedPatient.pk
                try:
                    appointmentData = Appointment.objects.filter(patientPres = selectedPatientID).order_by('date')
                    # serializedData = serialize('json', appointmentData) 
                except Appointment.DoesNotExist:
                    for singleappointmentData in appointmentData:
                        singleappointmentData = None    
                
                try:
                    prescriptionData = Prescription.objects.filter(patient_id = selectedPatientID).order_by('timestamp')
                    # prescriptionDataNew = Prescription.objects.get(prescribingPatient="Prem Suryan")

                    for presData in prescriptionData:
                        presmed = presData.medicine.all()
                        padMedSess = presData.MornAftNight.all()          

                        
                except Prescription.DoesNotExist :
                    for singleprescription in prescriptionData:
                        singleprescription = None

                        
                    else:
                        prescriptionData = None 

                doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
                
                context = {
                    "patients" : doctorSpecific,
                    "appointmentData" : appointmentData,
                    "prescriptionData" : prescriptionData, 
                    "presmed" : presmed,     
                    "padMedSess" : padMedSess              
                }
                
                response = render(request, "HealthCentre/timeline.html", context)
                return responseHeadersModifier(response)
                # return JsonResponse(data)
            except Patient.DoesNotExist:
                return JsonResponse({'error': 'Patient not found'}, status=404)
        

def deleteprescription(request, pk):
    # request.session['deleteAppointment'] = True
    delprescriptionobj = Prescription.objects.get(id=pk)
    delprescriptionobj.delete()
    doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
    records = doctor.doctorRecords.all()
    context = {
    "message" : "Successfully Logged In.",
    "isAuthenticated" : True,
    "user": records.order_by('-timestamp'),
    "prescriptions" : Prescription.objects.all().order_by('-timestamp')
        }
    # response = render(request, 'HealthCentre/prescriptionPortal.html', context)
    response = response = HttpResponseRedirect(reverse('login'))
    return responseHeadersModifier(response)

def onlineprescription(request):
    """Function to submit online prescription request to doctor."""

    # Calling session variables checker
    request = requestSessionInitializedChecker(request)

    # If the request method is get
    if request.method == "GET":

        # If the user is logged in
        if request.session['isLoggedIn']:

            # Portal only for patient prescription request submission, not for doctors
            if request.session['isDoctor']:

                # Storing message inside context variable
                # context = {
                #         "message":"Only for patients."
                # }

                # Editing response headers so as to ignore cached versions of pages
                # response = render(request, "HealthCentre/prescriptionPortal.html", context)
                response = render(request, "HealthCentre/userDoctorProfilePortal.html")
                return responseHeadersModifier(response)

            # If the user is a patient
            else:

                # Storing available doctors inside context variable
                context = {
                    "doctors" : Doctor.objects.all().order_by('specialization')
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request, "HealthCentre/prescriptionPortal.html", context)
                return responseHeadersModifier(response)

        # If the user is not logged in
        else:

            # Storing message inside context variable
            context = {
                    "message":"Please Login First."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/prescriptionPortal.html", context)
            return responseHeadersModifier(response)

    # If the user is posting the prescription request
    elif request.method == "POST":

        # Accepting only if the user is logged in
        if request.session['isLoggedIn']:

            # If the prescription is being submitted back by a doctor
            if request.session['isDoctor']:

                # Extracting information from post request
                prescriptionText = request.POST['prescription']

                # Updating the prescription and saving it
                prescription = Prescription.objects.get(pk = request.POST['prescriptionID'])
                prescription.prescriptionText = prescriptionText
                prescription.isCompleted = True
                prescription.isNew = True
                prescription.save()

                # Getting the records of the doctor
                records = Doctor.objects.get(emailHash = request.session['userEmail']).doctorRecords.all()

                # Storing required information inside context variable
                context = {
                    "user" : records,
                    "successPrescriptionMessage" : "Prescription Successfully Submitted."
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request, "HealthCentre/userDoctorProfilePortal.html", context)
                return responseHeadersModifier(response)

            # Else if the patient is submitting prescription request
            else:

                # Extracting information from post request and getting the corresponding doctor
                doctor = Doctor.objects.get(pk = request.POST["doctor"])
                symptoms = request.POST["symptoms"]

                # Saving the prescription under the concerned doctor
                prescription = Prescription(doctor = doctor, patient = Patient.objects.get(emailHash = request.session['userEmail']), symptoms = symptoms)
                prescription.save()

                # Storing information inside context variable
                context = {
                    "successPrescriptionMessage" : "Prescription Successfully Requested.",
                    "doctors"  : Doctor.objects.all().order_by('specialization')
                }

                # Editing response headers so as to ignore cached versions of pages
                # response = render(request, "HealthCentre/userDoctorProfilePortal.html", context)
                response = render(request, "HealthCentre/prescriptionPortal.html", context)
                return responseHeadersModifier(response)

        # Else if the user is not logged in
        else:

            # Storing information inside context variable
            context = {
                    "successPrescriptionMessage":"Please Login First.",
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/loginPortal.html", context)
            return responseHeadersModifier(response)

    # For any other method of access, returning a new blank online prescription page
    else:

        # Editing response headers so as to ignore cached versions of pages
        response = render(request, "HealthCentre/prescriptionPortal.html")
        return responseHeadersModifier(response)

def responseHeadersModifier(response):
    """Funtion to edit response headers so that no cached versions can be viewed. Returns the modified response."""
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response

def requestSessionInitializedChecker(request):
    """Function to initialize request sessions if they don't exist."""

    # Try except for KeyError
    try:
        # Checking if session variables exist
        if request.session['goToAppointmentsPage'] and request.session['CreatenewAppointment'] and request.session['isDoctor'] and request.session['isLoggedIn'] and request.session['userEmail'] and request.session['Name'] and request.session['numberNewPrescriptions'] and request.session['writeNewPrescription']:
            # Do nothing if they do exist
            pass
    except:
        # Initialize request variables if they don't exist
        request.session['isDoctor'] = ""
        request.session['isLoggedIn'] = False
        request.session['userEmail'] = ""
        request.session['Name'] = ""
        request.session['numberNewPrescriptions'] = ""
        request.session['writeNewPrescription'] = False
        request.session['CreatenewAppointment'] = False
        request.session['goToAppointmentsPage'] = False
        request.session['appointmentEdit'] = False
        request.session['patientMedEdit'] = False  
        request.session['medicineEdit'] = False 
        request.session['clickedOnAddRow'] = False
        request.session['medicineIsSelected'] = False
        request.session['sessionIsSelected'] = False
        # request.session['patientMedicineEdit'] = False
        dummyBoolean = False
 
    # Returning request
    return request

def whatsappNotification():
    
    while True:
        currentDateTime = datetime.now()
        currentHourObj = datetime.strftime(currentDateTime, "%H")
        currentMinuteObj = datetime.strftime(currentDateTime, "%M")
        currentDateObj = datetime.strftime(currentDateTime, "%d")
        currentMonthObj = datetime.strftime(currentDateTime, "%m")
        currentYearObj = datetime.strftime(currentDateTime, "%Y")

        strCurrenTimeObj = currentHourObj + currentMinuteObj
        strCurrentDateObj = currentDateObj + currentMonthObj + currentYearObj 

        plusthreehours=currentDateTime + timedelta(hours=3)
        hourObject = datetime.strftime(plusthreehours, "%H")
        minuteObject = datetime.strftime(plusthreehours, "%M")

        strTimeObject = hourObject + minuteObject
        datetimeObject = datetime.strptime(strTimeObject, "%H%M")
        datetimeObject = datetimeObject.time()

        currentDateTimeObj = datetime.strptime(strCurrenTimeObj, "%H%M")
        currentDateTimeObj = currentDateTimeObj.time()

        currentDate = datetime.strptime(strCurrentDateObj, "%d%m%Y")
        currentDate = currentDate.date()
        # getappointmentTime = Appointment.objects.get(time=datetimeObject)
        try:
            getAllAppointmentTime = Appointment.objects.filter(time=datetimeObject)

        except Appointment.DoesNotExist:
            for getappointmentTime in getAllAppointmentTime:
                getappointmentTime = None
        # time = models.TimeField(default=timezone.now)
        # getappointmentTime = Appointment.objects.get_or_(time=datetimeObject)
        for getappointmentTime in getAllAppointmentTime:
            if not getappointmentTime == None:
                AppointmentTime = getappointmentTime.time
                combinetime = datetime.combine(datetime.today(), AppointmentTime)
                AppointmentTime1 = datetime.strftime(combinetime, '%I:%M %p')
                AppointmentDate = getappointmentTime.date
                combinedate = datetime.combine(AppointmentDate, datetime.now().time())
                AppointmentDate1 = datetime.strftime(combinedate, "%b. %d, %Y")
                patientName=getappointmentTime.appointmentpatient
                doctorName = getappointmentTime.appointmentdoctor
                doctorDetail = Doctor.objects.get(name = doctorName)
                doctorNumber = doctorDetail.contactNumber
                patientDetail = Patient.objects.get(name=patientName)
                patientNumber = patientDetail.contactNumber
                if (AppointmentDate == currentDate):
                    whatsappApi(patientName, doctorName, patientNumber, AppointmentTime1, AppointmentDate1, doctorDetail.clinicName)
                    whatsappApiDoc(doctorName, doctorNumber, AppointmentTime1, AppointmentDate1, patientName, patientNumber) #AppointmentTime, AppointmentDate)
                    time.sleep(60)
                
        # while True:
        #     updateExcel()
        #     time.sleep(900)
def backgroundTask():
    thread = threading.Thread(target=whatsappNotification)
    thread.daemon = True
    thread.start()
backgroundTask()

def backgroundtastForUpdatingExcel():
    xlthread = threading.Thread(target= updateExcel)
    xlthread.daemon = True
    xlthread.start()
# backgroundtastForUpdatingExcel()
# generateqr : str

qrgen = ""
def catchgenqr(qrCode: str , asciiQR: str , attempt: int, urlCode: str):
    global qrgen
    qrgen = qrCode

def wpconnect():
    # docName = Settings.globalDocName
    global G_docName
    docName = G_docName
    openWhatsapp.wp(docName)
    # generateqr : str


def catchqrcode(request):
    
    # catchqr("data:image/png;base64,", "", 1, "2@242")
    global qrgen
    # global wpIsConnected
    qrdata = qrgen
    WhatsappIsConnected = settings.WP_IS_CONNECTED
    if not WhatsappIsConnected:
        request.session['wpStatus'] = "waiting for whatsapp QR code scan"
    else:
        request.session['wpStatus'] = "whatsapp is registered. Now login"
        
    context = {
        'qrdata': qrdata,
        
    }

    return render(request, 'HealthCentre/qrCode.html', context)

# wppstatus =""
# global creator
 
# creator = ""
# client =""
G_docName = ""
def whatsappBrowser(request):
    # global G_docName
    # docNameCommon = request.session['Name']
    # G_docName = docNameCommon
    
    backgroundtastForQrCode()
    response = HttpResponseRedirect(reverse('login'))
    return response

def whatsappStatus(request):
    creator = settings.GLOBAL_VAR
    global G_docName
    # docNameCommon = request.session['Name']
    # G_docName = docNameCommon
    
    # Settings.globalDocName = request.session['Name']
    if creator.state == 'CONNECTED' and (creator.session == G_docName):
        wppstatus = "Whatsapp is connected"
        request.session['wpStatus'] = True
    
    # elif creator.state == 'CONNECTED' and alreadyOpenStatus = 
    else:
        wppstatus = "Whatsapp is Disconnected"  
        request.session['wpStatus'] = False 
    data = {'wppStatus': wppstatus}

    return JsonResponse(data)


def generateqrcode():
    while True:
        global generateqr 
        
        # generateqr = catchqr()
        time.sleep(1)

def backgroundtastForQrCode():
    # openwp = openWhatsapp.wp()
    # openWhatsapp()
    
    qrthread = threading.Thread(target= wpconnect)
    qrthread.daemon = True
    qrthread.start()
    
# backgroundtastForQrCode()

def searchAppointments(request):
    
    if request.method == 'GET':
        response = response = HttpResponseRedirect(reverse('doctorappointmentsfalse'))
        return responseHeadersModifier(response)
    if request.method == "POST":
        searchDate = request.POST["searchByDate"]
        searchQuery = request.POST["searchQuery"]
        if searchQuery != '':


            searchFilterAppointments = Appointment.objects.filter((Q(appointmentpatient__icontains = searchQuery) |
                                                                Q(appointmentdoctor__icontains = searchQuery) |
                                                                Q(notes__icontains = searchQuery) |
                                                                Q(time__icontains = searchQuery) |
                                                                Q(subject__icontains = searchQuery)) & Q(appointmentdoctor = request.session['Name']))
            context = {
                'searchAppointmentPatients' : searchFilterAppointments.order_by('appointmentpatient'),
                'searchQuery' : searchQuery,
                
            }

            response = render(request, "HealthCentre/appointmentsPortal.html", context)
            return responseHeadersModifier(response)
        if  searchDate != '':
            searchFiterDate = Appointment.objects.filter(Q(date__icontains = searchDate))
            context ={
                'searchAppointmentPatients': searchFiterDate.order_by('appointmentpatient'),
                'searchDate' : searchDate
            }
            
            response = render(request, "HealthCentre/appointmentsPortal.html", context)
            return responseHeadersModifier(response)
        else:
            response = response = HttpResponseRedirect(reverse('doctorappointmentsfalse'))
            return responseHeadersModifier(response)


def searchPrescriptions(request):
    if request.method == 'GET':
        response = response = HttpResponseRedirect(reverse('login'))
        return responseHeadersModifier(response)
    
    if request.method == "POST":
        searchQuery = request.POST["searchQuery"]
        searchDate = request.POST["searchByDate"]
        if searchQuery != '':

            searchFilterPrescriptions = Prescription.objects.filter(Q(prescribingPatient__icontains = searchQuery) & Q(prescribingDoctor = request.session['Name'] ))

            context = {
                'searchPrescriptionPatients' : searchFilterPrescriptions.order_by('prescribingPatient'),
                'searchQuery' : searchQuery
            }

            response = render(request, "HealthCentre/prescriptionPortal.html", context)
            return responseHeadersModifier(response)
        
        if  searchDate != '':
            searchFiterDate = Prescription.objects.filter(Q(timestamp__icontains = searchDate))

            context ={
                'searchPrescriptionPatients': searchFiterDate.order_by('prescribingPatient'),
                'searchDate' : searchDate
            }

            response = render(request, "HealthCentre/prescriptionPortal.html", context)
            return responseHeadersModifier(response)
        else:
            response = response = HttpResponseRedirect(reverse('yourPrescriptions'))
            return responseHeadersModifier(response)



def countPrescriptionRows(request):
    if request.method == 'GET':
       clickedOnAddRow = False
       medicineIsSelected = False
       sessionIsSelected = False
       request.session['medicineIsSelected'] = False
       global rowCount
       rowCount += 1
       
       data = {
           'rowCOunt' : rowCount,
           'clickedOnAddRow' : clickedOnAddRow,
           'medicineIsSelected' : medicineIsSelected,
           'sessionIsSelected' : sessionIsSelected
       }
       return JsonResponse(data)

@csrf_exempt    
def presMedAjaxData(request):
    if request.method =='POST':
        # patObjId = Prescription.objects.get(pk=PatientSelected)
        medicineAjax = request.POST.get('medicineIsSelected')
        sessionAjax = request.POST.get('sessionIsSelected')

        data = {
            "medicineAjax":medicineAjax,
            "sessionAjax":sessionAjax
        }

        return JsonResponse(data)
        

def generatePDF(request):

    if request.method == "GET":
        global rowCount
        global additionalRowData
        # pyautogui.hotkey('ctrl', 'p')
        # prescriptionTemplate = (get_template("HealthCentre/NewPrescription.html"))
        addedData = request.session.get('additionalSession_data', {})
        addedData2 = request.session.get('additionalMedicine_data', {})
        addedData.update(addedData2)
        allAddData = {}
        allAddData.update(additionalRowData)
        newRowdata = allAddData
        request.session['newRowData'] = newRowdata
        data = request.session.get('pdf_data' , {}) #+ 'pdf_sess_data' + 'pdf_med_data'
        data1 = request.session.get('pdf_sess_data' , {})
        data2 = request.session.get('pdf_med_data' , {})
        data3 = request.session.get('pdf_days_data' , {})
        data4 = request.session.get('newRowData', {})
        imagePath = request.session.get('imagePath')
        data.update(data1)
        data.update(data2)
        data.update(data3)
        data.update(data4)
        data.update(imagePath)
        patientName = data.get("patientName")
        currentDate = date.today() 
        renderedHTML = render_to_string(r'HealthCentre\NewPrescriptionforRendering.html',data, request)
        
        additionalRows = ''
        for i in range(int(rowCount)):
            additionalRows += f'''<tr class = "prescription-row"><td><h6 name="beforeAfter" id = "beforeAfter" class="form-control">{ data4.get(f'MedBefAft{i+1}')} </h6></td>
            <td>
              <h6 name="SelectedMed" id = "SelectedMed" class="form-control">{ data4.get(f'SelectedMed{i+1}') }</h6>
            </td>
            <td>
              <h6 name="morning" id = "morning" class="form-group">{ data4.get(f'MedMorn{i+1}') }</h6>
            </td>
            <td>
              <h6 name="afternoon" id = "afternoon" class="form-group">{ data4.get(f'medAft{i+1}') }</h6>
            </td>
            <td>
              <h6 name="night" id = "night" class="form-control">{ data4.get(f'medNight{i+1}')}</h6>
            </td>
            <td>
            </td>
          </tr>'''
        renderedHTML = renderedHTML.replace("<tbody>\n          </tbody>\n", f'<tbody>{additionalRows}</tbody>')
        renderedfile = pdfkit.from_string(renderedHTML,f'E:\Adhithya\dental-software\DentalCareDev\prescPDF\{patientName}_{currentDate}.pdf', options={"enable-local-file-access": None})
        rowCount = 0
    return HttpResponseRedirect(reverse("doctorprofile"))

def sendPdfinWhatsapp(wpnumber, docName, patientname, curDate):

    curPath = os.getcwd() #"D:\prescPDF"
    pdfPath = os.path.join(curPath, "prescPDF")
    allFilesInPath = os.listdir(pdfPath)
    # filesInPath = [allfiles for allfiles in allFilesInPath if allfiles.lower().startswith('prescPDF')]
    # if filesInPath:
    doctorName = docName
    PdfFilesInPath = [file for file in allFilesInPath if file.lower().endswith('.pdf')]
    
    if not PdfFilesInPath:
        pass
    else:
        pdfFullPaths = [os.path.join(pdfPath, pdfFile)  for pdfFile in PdfFilesInPath] 
        latestPdf = max(pdfFullPaths, key=os.path.getmtime)
    
    whatsappMedia(wpnumber, latestPdf, doctorName, patientname, curDate)

def patMed(request):
    request.session['patientMedEdit'] = False
    if request.method == "GET":
        request.session['patientMedicineEdit'] = False
        # patient = Patient.objects.get(id)
        doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
        context = {
            "editPat" : doctorSpecific,
            "editMedicine" : Medicine.objects.all().order_by('medicinename'),
            }

        response = render(request, "HealthCentre/medicinePatientPortal.html", context)
        return responseHeadersModifier(response)

def editPatientMed(request,pk):
    request.session['patientMedEdit'] = True
    if request.method == "POST": 
        request.session['patientMedEdit'] = False  
        patientObject = Patient.objects.get(id=pk)
        patientObject.name = request.POST['userFirstNam']
        patientObject.address = request.POST['userAddress']
        patientObject.contactNumber = request.POST['userContactNo']
        patientObject.email = request.POST['userEmail']
        patientObject.rollNumber = request.POST['userRollNo']
        patientObject.passwordHash = request.POST['userPassword']
        doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
        patientObject.save()   
        context = {
               "editPat" : doctorSpecific,
               "editMedicine" : Medicine.objects.all().order_by('medicinename'),
             }
        response = HttpResponseRedirect(reverse('patMed'))
        return responseHeadersModifier(response)

    patientObj = Patient.objects.get(id=pk)
    patientName = patientObj.name     
    patientAddr = patientObj.address
    patientContact = patientObj.contactNumber
    patientEmail = patientObj.email
    patientRoll = patientObj.rollNumber
    patientSex = patientObj.passwordHash
    doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')     
    context= {
        "userFirstNam" :patientName,
        "userAddress" : patientAddr,
        "userContactNo" : patientContact,
        "userEmail" : patientEmail,
        "userRollNo" : patientRoll,
        "editPat" : doctorSpecific,
        "patientSex" : patientSex,
        "editMedicine" : Medicine.objects.all(),

    }
    response = render(request, "HealthCentre/medicinePatientPortal.html", context)
    return responseHeadersModifier(response)

def deletepatientDetails(request, pk):
    request.session['deletepatientDetails'] = True
    delpatObj = Patient.objects.get(id=pk)
    delpatObj.delete()
    doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
    context = {
    
    "delpatObj" : delpatObj,
     "editPat" : doctorSpecific,
    "editMedicine" : Medicine.objects.all().order_by('medicinename'),
        }
    response = response = HttpResponseRedirect(reverse('patMed'))
    return responseHeadersModifier(response)


def medicineEdit(request,pk):
    request.session['medicineEdit'] = True
    if request.method == "POST": 
        request.session['medicineEdit'] = False  
        medicineObject = Medicine.objects.get(id=pk)
        medicineObject.medicinename = request.POST['patientMed']
        medicineObject.beforeafter = request.POST['beforeAfter']
        medicineObject.morning = '1'
        medicineObject.afternoon = '1'
        medicineObject.night = '1'
        doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
        medicineObject.save()   
        context = {
                "editPat" : doctorSpecific,
               "editMedicine" : Medicine.objects.all().order_by('medicinename'),
             }
        response = HttpResponseRedirect(reverse('patMed'))
        return responseHeadersModifier(response)

    medicineObj = Medicine.objects.get(id=pk)
    medicineName = medicineObj.medicinename    
    befAftr = medicineObj.beforeafter
    doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
    context= {
       "medicineName" : medicineName,
       "befAftr" : befAftr,
        "editPat" : doctorSpecific,
        "editMedicine" : Medicine.objects.all().order_by('medicinename'),

    }
    response = render(request, "HealthCentre/medicinePatientPortal.html", context)
    return responseHeadersModifier(response)

def deletemedicineDetails(request, pk):
    request.session['deletemedicineDetails'] = True
    delmedObj = Medicine.objects.get(id=pk)
    delmedObj.delete()
    doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
    
    context = {
    
    "delmedObj" : delmedObj,
     "editPat" : doctorSpecific,
    "editMedicine" : Medicine.objects.all().order_by('medicinename'),
        }
    response = response = HttpResponseRedirect(reverse('patMed'))
    return responseHeadersModifier(response)

def searchPatients(request):

    if request.method == 'GET':
        response = response = HttpResponseRedirect(reverse('patMed'))
        return responseHeadersModifier(response)
    if request.method == 'POST':
        searchQuery = request.POST["searchQuery"]
        doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
        if searchQuery != '':

            searchFilterPatients = Patient.objects.filter((Q(name__icontains = searchQuery) |
                                                                Q(address__icontains = searchQuery) |
                                                                Q(contactNumber__icontains = searchQuery) |
                                                                Q(email__icontains = searchQuery) |
                                                                Q(rollNumber__icontains = searchQuery)) &
                                                                Q(doctorname = request.session['Name']))
            context = {
                'editPat' : searchFilterPatients.order_by('name'),
                "editMedicine" : Medicine.objects.all(),
                'searchQuery' : searchQuery
            }

        
            response = render(request, "HealthCentre/medicinePatientPortal.html", context)
            return responseHeadersModifier(response)  
        else:
            response = response = HttpResponseRedirect(reverse('patMed'))
            return responseHeadersModifier(response)

def searchMedicine(request):
    if request.method == 'GET':
        response = response = HttpResponseRedirect(reverse('patMed'))
        return responseHeadersModifier(response)
    if request.method == 'POST':
        searchQuery = request.POST["searchQuery"]
        doctorSpecific = Patient.objects.filter(doctorname = request.session['Name']).order_by('name')
        if searchQuery != '':

            searchFilterMedicine = Medicine.objects.filter(Q(medicinename__icontains = searchQuery) |
                                                                Q(beforeafter__icontains = searchQuery))
            context = {
                'editMedicine' : searchFilterMedicine.order_by('medicinename'),
                "editPat" : doctorSpecific,
                'searchQuery' : searchQuery
            }
            response = render(request, "HealthCentre/medicinePatientPortal.html", context)
            return responseHeadersModifier(response)
        else:
            response = HttpResponseRedirect(reverse('patMed'))
            return responseHeadersModifier(response)        
    









# def updateGoogleSheets():
#     excelFilePath = 'D:\Dental-Software\database_tables.xlsx'
#     GsWorkbook = openpyxl.load_workbook(excelFilePath)
#     GSWorksheet = GsWorkbook.active
#     GsheetUrl = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSnNGw6Mh3W9nuNkUku96aWP04fJMyTqMwJrOQsgtAAbCkdcAcafs_SH85Ve9IluvjXdulA8HZnPTXy/pubhtml'

#     for row in GSWorksheet.iter_rows(values_only = True):
#         data = ','.join(map(str, row))
#         response = requests.get(f'{GsheetUrl}?gid=0&single=true&output=csv&exportformat=csv&range=A1', data=data)
    
#         if response.status_code == 200:
#             print(f'Successfully updated: {data}')
#         else:
#             print(f'Failed to update: {data}')